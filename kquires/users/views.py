from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.db.models import QuerySet
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.views.generic import DetailView, RedirectView, UpdateView, ListView, TemplateView, CreateView, View
from django.contrib.auth.hashers import check_password, make_password
from kquires.users.models import User
from django.shortcuts import render, redirect
from django.contrib import messages
from .forms import UserForm, UserUploadForm
from django.shortcuts import get_object_or_404
from django.urls import reverse_lazy
from django.http import JsonResponse
import csv
from openpyxl import load_workbook
from datetime import date
from ..departments.models import Department
from django.http import HttpResponse, HttpRequest
from django.db.models import Q
import json
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.generic.edit import FormView
from allauth import app_settings as allauth_app_settings
from allauth.account import app_settings
from allauth.decorators import rate_limit
from django.views.decorators.debug import sensitive_post_parameters
from django.views.decorators.cache import never_cache
from allauth.utils import get_form_class
from allauth.core.exceptions import ImmediateHttpResponse
from django.contrib.sites.shortcuts import get_current_site
from allauth.account.internal import flows
from django.core.validators import validate_email
from django.forms import ValidationError
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import DeleteView
from .models import User
from allauth.account.internal.decorators import (
    login_not_required,
)
from allauth.account.forms import (

    SignupForm,

)

from allauth.account.mixins import (
    AjaxCapableProcessFormViewMixin,
    CloseableSignupMixin,
    NextRedirectMixin,
    RedirectAuthenticatedUserMixin,
    _ajax_response,
)
from django.http import HttpResponseRedirect

@method_decorator(csrf_exempt, name='dispatch')
class LoginView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Check if request is JSON
            if request.content_type == "application/json":
                data = json.loads(request.body)
            else:
                # Handle form-encoded data
                data = request.POST

            username = data.get("username")  # Can be email or employee_id
            password = data.get("password")

            if not username or not password:
                return JsonResponse({"error": "Username and password are required"}, status=400)

            if '@' in username:
                user = User.objects.filter(email=username).first()
                if not user:
                    return JsonResponse({"error": "Invalid Email"}, status=400)
            else:
                user = User.objects.filter(employee_id=username).first()
                if not user:
                    return JsonResponse({"error": "Invalid Employee ID"}, status=400)
                # Use the associated email for authentication
                username = user.email  # Replace employee_id with email for authentication

            # Authenticate using email (USERNAME_FIELD = "email")
            if not user.is_superuser and user.status != 'True':
                return JsonResponse({"error": "User Status is Inactive"}, status=400)
            user = authenticate(request, username=username, password=password)

            if user:
                login(request, user)
                return JsonResponse({
                    "message": "Login successful",
                    "user": {
                        "email": user.email,
                        "employee_id": user.employee_id,
                        "name": user.name,
                        "role": user.role
                    }
                })
            else:
                return JsonResponse({"error": "Invalid credentials"}, status=400)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format"}, status=400)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


class UserProfileView(LoginRequiredMixin, DetailView):
    model = User
    template_name = 'users/profile.html'
    context_object_name = 'user'

    def get_object(self, queryset=None):
        return self.request.user

class UserPasswordResetView(LoginRequiredMixin, TemplateView):
    template_name = 'users/password_reset.html'

    def post(self, request, *args, **kwargs):
        # Get the authenticated user
        user = request.user

        # Get the form data
        old_password = request.POST.get("old_password")
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        # Check if old password is correct
        if not user.check_password(old_password):
            messages.error(request, "Old password is incorrect.")
            return redirect("users:password_reset")

        # Check if new password and confirm password match
        if new_password != confirm_password:
            messages.error(request, "New password and confirm password do not match.")
            return redirect("users:password_reset")

        # Update the password
        user.password = make_password(new_password)
        user.save()

        messages.success(request, "Password updated successfully.")
        return redirect("users:password_reset")




class UserDetailView(LoginRequiredMixin, DetailView):
    model = User
    slug_field = "id"
    slug_url_kwarg = "id"


user_detail_view = UserDetailView.as_view()


class UserUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = User
    fields = ["name"]
    success_message = _("Information successfully updated")

    def get_success_url(self) -> str:
        assert self.request.user.is_authenticated  # type guard
        return self.request.user.get_absolute_url()

    def get_object(self, queryset: QuerySet | None=None) -> User:
        assert self.request.user.is_authenticated  # type guard
        return self.request.user


user_update_view = UserUpdateView.as_view()


class UserRedirectView(LoginRequiredMixin, RedirectView):
    permanent = False

    def get_redirect_url(self) -> str:
        return reverse("users:detail", kwargs={"pk": self.request.user.pk})


user_redirect_view = UserRedirectView.as_view()



class UserListView(LoginRequiredMixin, ListView):
    model = User
    template_name = 'users/list.html'
    context_object_name = 'users'
    paginate_by = 10

    def get_queryset(self):
        queryset = User.objects.filter(is_superuser=False).order_by('-updated_at')
        query = self.request.GET.get("q", "").strip()
        
        if query:
            queryset = queryset.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(job_title__icontains=query) |
            Q(employee_id__icontains=query)
        )
        return queryset

    def dispatch(self, request, *args, **kwargs):
        user = self.request.user

        # Redirect to the dashboard home page
        if not user.is_admin:
            return redirect('dashboard:index')  # Redirect to home page (dashboard/index.html)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['today_date'] = date.today().strftime("%Y-%m-%d")
        context["departments"] = Department.objects.all()
        return context

    def get_paginate_by(self, queryset: HttpRequest):
        page_size = self.request.GET.get('page_size', 10)
        try:
            return int(page_size)
        except ValueError:
            return 10

    def render_to_response(self, context, **response_kwargs):
        request = self.request
        if request.headers.get("HX-Request") == "true":
            return render(request, "users/partials/table.html", context)
        return super().render_to_response(context, **response_kwargs)




class UserCreateOrUpdateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = User
    form_class = UserForm
    template_name = 'users/list.html'
    success_url = reverse_lazy('users:list')

    def get_object(self, queryset=None):
        # Get the user_id from either POST or GET
        user_id = self.request.POST.get('id') or self.request.GET.get('id')
        if user_id:
            return get_object_or_404(User, id=user_id)
        return None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["user"] = self.get_object()
        return context

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        user = self.get_object()
        if user:
            kwargs['instance'] = user
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        self.object = form.save(commit=False)
        if not self.object.pk:  # If this is a new user
            self.object.created_by = self.request.user
        self.object.save()
        action = "updated" if self.object.pk else "created"
        user =  User.objects.get(id=self.request.user.id)
        user.log(action, f"User successfully {action}.")
        messages.success(self.request, f"User successfully {action}.")
        return super().form_valid(form)

    def form_invalid(self, form):
        error_messages = [
            f"{form.fields[field].label if field != '__all__' else 'Error'}: {error}"
            for field, errors in form.errors.items()
            for error in errors
        ]
        messages.error(self.request, "\n".join(error_messages))
        return HttpResponseRedirect(self.success_url)


def user_detail_api(request, id):
    user = User.objects.get(id=id)
    department_name = user.department.name if user.department else None
    return JsonResponse({
        'id': user.id,
        'employee_id': user.employee_id,
        'name': user.name,
        'email': user.email,
        'phone_number': user.phone_number,
        'mobile_number': user.mobile_number,
        'job_title': user.job_title,
        'status': user.status,
        'registration_date': user.registration_date,
        'department': department_name,
        'role': user.role,
        'email_enabled': user.email_enabled,
    })



class UserUploadView(LoginRequiredMixin, View):
    template_name = "users/list.html"

    def get(self, request):
        form = UserUploadForm()
        return render(request, self.template_name, {"form": form})

    def post(self, request):
        form = UserUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES["file"]
            try:
                # Read and process the file
                if file.name.endswith(".csv"):
                    with open(file, mode="r", encoding="utf-8") as f:
                        reader = csv.DictReader(f)  # Use DictReader to handle rows as dictionaries
                        data = [row for row in reader]
                elif file.name.endswith(".xlsx"):
                    workbook = load_workbook(file)
                    sheet = workbook.active
                    data = []
                    headers = [cell.value for cell in sheet[1]]  # First row as headers
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                        data.append(dict(zip(headers, row)))  # Map headers to row values
                else:
                    messages.error(request, "Unsupported file format. Please upload a CSV or Excel file.")
                    return redirect("users:list")

                # Validate required columns
                required_columns = ["email", "name", "job_title", "department", "role"]
                for column in required_columns:
                    if column not in data[0]:
                        messages.error(request, f"Missing required column: {column}")
                        return redirect("users:list")

                # Iterate through rows and create users
                for row in data:
                    if not User.objects.filter(email=row["email"]).exists():  # Avoid duplicates
                        if row.get("department"):
                            department, _ = Department.objects.get_or_create(name=row["department"])
                        user = User.objects.create(
                            employee_id=row["employee_id"],
                            email=row["email"],
                            name=row["name"],
                            # phone_number=row.get("phone_number", ""),
                            # mobile_number=row.get("mobile_number", ""),
                            job_title=row.get("job_title", ""),
                            department=department,
                            registration_date=row.get("registration_date", date.today()),
                            role=row.get("role", ""),
                            status = 'True',
                            is_active = True,
                            email_enabled = 'on',
                        )
                        password = str(row.get("password", "Password123!"))
                        role = row.get("role").lower()
                        user.set_password(password)
                        if hasattr(user, f'is_{role}'):
                            setattr(user, f'is_{role}', True)
                        else:
                            setattr(user, f'is_staff', True)
                        user.save()
                user =  User.objects.get(id=self.request.user.id)
                user.log('uploaded', f"Users file uploaded.")
                messages.success(request, "Users uploaded successfully!")
                return redirect("users:list")
            except Exception as e:
                messages.error(request, f"Error processing file: {str(e)}")
                return redirect("users:list")
        else:
            messages.error(request, "Invalid form submission.")
        return render(request, self.template_name, {"form": form})

def export_csv(request):
    # Define the response as a CSV file
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users.csv"'

    # Create a CSV writer
    writer = csv.writer(response)

    writer.writerow([
        'ID', 'Name', 'Email', 'Employee ID',  'Phone Number', 'Mobile Number', 'Job Title',
        'Email Enabled', 'Registration Date', 'Department', 'Role',
        'Status', 'Is Admin', 'Created At'
    ])

    # Write data rows
    users = User.objects.all()  # Fetch data
    for user in users:
        writer.writerow([
            user.id,
            user.name,
            user.email,
            user.employee_id,
            user.phone_number,
            user.mobile_number,
            user.job_title,
            user.email_enabled,
            user.registration_date,
            user.department.name if user.department else "N/A",
            user.role,
            user.status,
            user.is_admin,
            user.created_at
        ])

    return response



class SignupView(
    RedirectAuthenticatedUserMixin,
    CloseableSignupMixin,
    NextRedirectMixin,
    AjaxCapableProcessFormViewMixin,
    FormView,
):
    template_name = "account/signup." + app_settings.TEMPLATE_EXTENSION
    form_class = SignupForm

    @method_decorator(rate_limit(action="signup"))
    @method_decorator(login_not_required)
    @sensitive_post_parameters()
    @method_decorator(never_cache)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_form_class(self):
        return get_form_class(app_settings.FORMS, "signup", self.form_class)

    def form_valid(self, form):
        self.user, resp = form.try_save(self.request)
        if resp:
            return resp
        try:
            redirect_url = self.get_success_url()
            return flows.signup.complete_signup(
                self.request,
                user=self.user,
                redirect_url=redirect_url,
                by_passkey=form.by_passkey,
            )
        except ImmediateHttpResponse as e:
            return e.response

    def get_context_data(self, **kwargs):
        ret = super().get_context_data(**kwargs)
        passkey_signup_enabled = False
        if allauth_app_settings.MFA_ENABLED:
            from allauth.mfa import app_settings as mfa_settings

            passkey_signup_enabled = mfa_settings.PASSKEY_SIGNUP_ENABLED
        form = ret["form"]
        email = self.request.session.get("account_verified_email")
        if email:
            email_keys = ["email"]
            if "email2" in app_settings.SIGNUP_FIELDS:
                email_keys.append("email2")
            for email_key in email_keys:
                form.fields[email_key].initial = email
        login_url = self.passthrough_next_url(reverse("account_login"))
        signup_url = self.passthrough_next_url(reverse("account_signup"))
        signup_by_passkey_url = None
        if passkey_signup_enabled:
            signup_by_passkey_url = self.passthrough_next_url(
                reverse("account_signup_by_passkey")
            )
        site = get_current_site(self.request)
        ret.update(
            {
                "login_url": login_url,
                "signup_url": signup_url,
                "signup_by_passkey_url": signup_by_passkey_url,
                "site": site,
                "SOCIALACCOUNT_ENABLED": allauth_app_settings.SOCIALACCOUNT_ENABLED,
                "SOCIALACCOUNT_ONLY": allauth_app_settings.SOCIALACCOUNT_ONLY,
                "PASSKEY_SIGNUP_ENABLED": passkey_signup_enabled,
            }
        )
        return ret

    def get_initial(self):
        initial = super().get_initial()
        email = self.request.GET.get("email")
        if email:
            try:
                validate_email(email)
            except ValidationError:
                return initial
            initial["email"] = email
            if "email2" in app_settings.SIGNUP_FIELDS:
                initial["email2"] = email
        return initial


signup = SignupView.as_view()





class UserDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = User
    success_url = reverse_lazy('users:list')  # Adjust to your users list URL name
    pk_url_kwarg = 'id'  # Ensure your URL pattern matches this

    def get_object(self, queryset=None):
        return get_object_or_404(User, id=self.kwargs.get(self.pk_url_kwarg))

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        request.user.log('deleted', f"User '{obj.name}' (email: {obj.email}) successfully deleted.")
        messages.success(request, _("User successfully deleted."))
        return super().delete(request, *args, **kwargs)

    def test_func(self):
        # Only allow deletion if the user is an admin
        return self.request.user.is_authenticated and self.request.user.is_admin